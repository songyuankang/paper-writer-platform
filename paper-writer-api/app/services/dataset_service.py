"""Independent, versioned research dataset storage.

Dataset metadata lives under ``data/datasets/<dataset_id>/metadata.json`` while
rows are stored per immutable version as controlled CSV files.  This deliberately
keeps large research data out of ``draft.json`` and leaves material_service's
text-oriented upload workflow untouched.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import shutil
import statistics
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from app.config import Settings

MAX_PREVIEW_ROWS = 200
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
_BOOL_VALUES = {"true", "false", "yes", "no", "y", "n", "是", "否", "1", "0"}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _clean(value: object, limit: int = 1000) -> str:
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip()[:limit]


def _number(value: object) -> float | None:
    try:
        text = _clean(value).replace(",", "")
        if not text:
            return None
        parsed = float(text)
        return parsed if math.isfinite(parsed) else None
    except (ValueError, TypeError):
        return None


def _datetime_like(value: object) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    text = _clean(value)
    if not text:
        return False
    if not re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}(?:[ T].*)?", text):
        return False
    try:
        datetime.fromisoformat(text.replace("/", "-"))
        return True
    except ValueError:
        return False


def _detect_csv_text(raw: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8-replace"


def _headers(values: list[object]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = _clean(value, 120) or f"列{index + 1}"
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return result


def _fingerprint(headers: list[str], rows: list[dict[str, str]]) -> str:
    raw = json.dumps({"headers": headers, "rows": rows}, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _column_profile(name: str, values: list[str]) -> dict[str, Any]:
    present = [item for item in values if _clean(item)]
    missing_count = len(values) - len(present)
    unique_count = len(set(present))
    numeric = [_number(item) for item in present]
    numeric_values = [item for item in numeric if item is not None]
    numeric_ratio = len(numeric_values) / len(present) if present else 0
    bool_ratio = sum(_clean(item).lower() in _BOOL_VALUES for item in present) / len(present) if present else 0
    date_ratio = sum(_datetime_like(item) for item in present) / len(present) if present else 0
    warnings: list[str] = []
    if present and numeric_ratio == 1:
        kind = "numeric"
    elif present and bool_ratio == 1:
        kind = "boolean"
    elif present and date_ratio == 1:
        kind = "datetime"
    elif present and numeric_ratio >= .7:
        kind = "numeric"
        warnings.append("该列多数值为数字，但包含非数字内容。")
    elif unique_count <= min(50, max(2, int(len(present) * .5))):
        kind = "categorical"
    else:
        kind = "text"
    stats: dict[str, float] | None = None
    if kind == "numeric" and numeric_values:
        stats = {
            "mean": sum(numeric_values) / len(numeric_values),
            "median": float(statistics.median(numeric_values)),
            "min": min(numeric_values),
            "max": max(numeric_values),
            "std": float(statistics.stdev(numeric_values)) if len(numeric_values) > 1 else 0.0,
        }
    if missing_count:
        warnings.append(f"缺失值 {missing_count} 个。")
    return {
        "name": name, "type": kind, "nullable": missing_count > 0,
        "unique_count": unique_count, "missing_count": missing_count,
        "stats": stats, "warnings": warnings,
    }


def _quality(headers: list[str], rows: list[dict[str, str]]) -> dict[str, Any]:
    columns = [_column_profile(header, [row.get(header, "") for row in rows]) for header in headers]
    duplicate_rows = len(rows) - len({json.dumps(row, ensure_ascii=False, sort_keys=True) for row in rows})
    warnings = [f"发现 {duplicate_rows} 行重复记录。"] if duplicate_rows else []
    for column in columns:
        warnings.extend([f"{column['name']}：{warning}" for warning in column["warnings"] if "缺失值" not in warning])
    return {
        "sample_size": len(rows), "variable_count": len(headers), "duplicate_rows": duplicate_rows,
        "columns": columns, "warnings": warnings,
    }


class DatasetService:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.root = settings.db_path.parent / "datasets"
        self.staging = settings.db_path.parent / "dataset_staging"
        self.root.mkdir(parents=True, exist_ok=True)
        self.staging.mkdir(parents=True, exist_ok=True)

    def _dir(self, dataset_id: str) -> Path:
        safe = re.sub(r"[^a-zA-Z0-9_-]", "", dataset_id)
        if not safe or safe != dataset_id:
            raise ValueError("数据集 ID 无效")
        return self.root / safe

    def _metadata_path(self, dataset_id: str) -> Path:
        return self._dir(dataset_id) / "metadata.json"

    def _load_metadata(self, dataset_id: str) -> dict[str, Any]:
        path = self._metadata_path(dataset_id)
        if not path.is_file():
            raise ValueError("未找到数据集")
        return json.loads(path.read_text(encoding="utf-8"))

    def _save_metadata(self, metadata: dict[str, Any]) -> None:
        directory = self._dir(str(metadata["id"]))
        directory.mkdir(parents=True, exist_ok=True)
        self._metadata_path(str(metadata["id"])).write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")

    def list_datasets(self, task_id: str | None = None) -> list[dict[str, Any]]:
        results: list[dict[str, Any]] = []
        for directory in self.root.iterdir() if self.root.exists() else []:
            path = directory / "metadata.json"
            if not path.is_file():
                continue
            try:
                metadata = json.loads(path.read_text(encoding="utf-8"))
                if task_id and task_id not in (metadata.get("task_ids") or []):
                    continue
                results.append(self.summary(metadata))
            except (ValueError, json.JSONDecodeError):
                continue
        return sorted(results, key=lambda item: item.get("updated_at", ""), reverse=True)

    @staticmethod
    def summary(metadata: dict[str, Any]) -> dict[str, Any]:
        latest = (metadata.get("versions") or [])[-1] if metadata.get("versions") else {}
        return {
            "id": metadata.get("id"), "name": metadata.get("name"), "description": metadata.get("description", ""),
            "source_type": metadata.get("source_type"), "created_at": metadata.get("created_at"), "updated_at": metadata.get("updated_at"),
            "latest_version": metadata.get("latest_version"), "task_ids": metadata.get("task_ids") or [],
            "row_count": latest.get("row_count", 0), "variable_count": len(latest.get("schema") or []),
            "latest_source": latest.get("source") or {},
        }

    def get_dataset(self, dataset_id: str) -> dict[str, Any]:
        metadata = self._load_metadata(dataset_id)
        return {**metadata, "summary": self.summary(metadata)}

    def versions(self, dataset_id: str) -> list[dict[str, Any]]:
        return list(self._load_metadata(dataset_id).get("versions") or [])

    def _read_rows(self, dataset_id: str, version: int) -> tuple[dict[str, Any], list[dict[str, str]]]:
        metadata = self._load_metadata(dataset_id)
        info = next((item for item in metadata.get("versions") or [] if int(item.get("version", 0)) == int(version)), None)
        if info is None:
            raise ValueError("未找到数据集版本")
        target = self._dir(dataset_id) / str(info["rows_path"])
        if not target.is_file():
            raise ValueError("数据集版本文件不存在")
        with target.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [{key: _clean(value) for key, value in row.items()} for row in csv.DictReader(handle)]
        return info, rows

    def get_version(self, dataset_id: str, version: int | None = None, include_rows: bool = False) -> dict[str, Any]:
        metadata = self._load_metadata(dataset_id)
        target_version = int(version or metadata.get("latest_version") or 0)
        info, rows = self._read_rows(dataset_id, target_version)
        result = {**info, "dataset_id": dataset_id, "dataset_name": metadata.get("name"), "source_type": "research_dataset"}
        if include_rows:
            result["rows"] = rows
        return result

    def preview(self, dataset_id: str, version: int, limit: int = 50, offset: int = 0) -> dict[str, Any]:
        info, rows = self._read_rows(dataset_id, version)
        safe_limit = min(max(1, limit), MAX_PREVIEW_ROWS)
        safe_offset = max(0, offset)
        return {
            "dataset_id": dataset_id, "version": int(version), "schema": info.get("schema") or [],
            "quality": info.get("quality") or {}, "rows": rows[safe_offset:safe_offset + safe_limit],
            "row_count": info.get("row_count", len(rows)), "limit": safe_limit, "offset": safe_offset,
            "has_more": safe_offset + safe_limit < len(rows),
        }

    def inspect_upload(self, filename: str, raw: bytes) -> dict[str, Any]:
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_EXTENSIONS:
            raise ValueError("研究数据仅支持 CSV 或 XLSX")
        if suffix == ".csv":
            return {"source_type": "csv", "sheets": [], "requires_sheet_selection": False}
        import openpyxl
        workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheets = list(workbook.sheetnames)
        workbook.close()
        return {"source_type": "xlsx", "sheets": sheets, "requires_sheet_selection": len(sheets) > 1}

    def stage_upload(self, filename: str, raw: bytes) -> dict[str, Any]:
        info = self.inspect_upload(filename, raw)
        token = uuid.uuid4().hex
        suffix = Path(filename).suffix.lower()
        path = self.staging / f"{token}{suffix}"
        path.write_bytes(raw)
        return {**info, "import_token": token, "filename": Path(filename).name}

    def _read_source(self, filename: str, raw: bytes, sheet: str | None = None) -> tuple[list[str], list[dict[str, str]], dict[str, Any]]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            text, encoding = _detect_csv_text(raw)
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            records = list(csv.reader(io.StringIO(text), dialect))
            source = {"filename": Path(filename).name, "extension": "csv", "encoding": encoding}
        elif suffix == ".xlsx":
            import openpyxl
            workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            target_sheet = sheet or workbook.sheetnames[0]
            if target_sheet not in workbook.sheetnames:
                workbook.close()
                raise ValueError("所选工作表不存在")
            records = [list(row) for row in workbook[target_sheet].iter_rows(values_only=True)]
            workbook.close()
            source = {"filename": Path(filename).name, "extension": "xlsx", "sheet": target_sheet}
        else:
            raise ValueError("研究数据仅支持 CSV 或 XLSX")
        records = [row for row in records if any(_clean(value) for value in row)]
        if len(records) < 2:
            raise ValueError("数据文件至少需要表头和一行数据")
        headers = _headers(records[0])
        rows = []
        for raw_row in records[1:]:
            rows.append({headers[index]: _clean(raw_row[index] if index < len(raw_row) else "") for index in range(len(headers))})
        return headers, rows, source

    def import_data(self, *, filename: str, raw: bytes, name: str = "", description: str = "", dataset_id: str | None = None, sheet: str | None = None, task_id: str | None = None) -> dict[str, Any]:
        headers, rows, source = self._read_source(filename, raw, sheet)
        fingerprint = _fingerprint(headers, rows)
        quality = _quality(headers, rows)
        if dataset_id:
            metadata = self._load_metadata(dataset_id)
            latest = (metadata.get("versions") or [])[-1] if metadata.get("versions") else None
            if latest and latest.get("fingerprint") == fingerprint:
                if task_id and task_id not in metadata.get("task_ids", []):
                    metadata.setdefault("task_ids", []).append(task_id)
                    metadata["updated_at"] = _now()
                    self._save_metadata(metadata)
                return {**latest, "dataset_id": dataset_id, "dataset_name": metadata["name"], "summary": self.summary(metadata), "deduplicated": True}
            version_number = int(metadata.get("latest_version") or 0) + 1
        else:
            dataset_id = f"ds_{uuid.uuid4().hex[:16]}"
            created_at = _now()
            metadata = {
                "id": dataset_id, "name": _clean(name, 120) or Path(filename).stem[:120] or "研究数据集",
                "description": _clean(description, 500), "source_type": source["extension"],
                "created_at": created_at, "updated_at": created_at, "latest_version": 0,
                "task_ids": [], "versions": [],
            }
            version_number = 1
        dataset_dir = self._dir(dataset_id)
        versions_dir = dataset_dir / "versions"
        source_dir = dataset_dir / "source"
        versions_dir.mkdir(parents=True, exist_ok=True)
        source_dir.mkdir(parents=True, exist_ok=True)
        rows_rel = f"versions/v{version_number}.csv"
        rows_path = dataset_dir / rows_rel
        with rows_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        source_rel = f"source/v{version_number}_{Path(filename).name}"
        (dataset_dir / source_rel).write_bytes(raw)
        version = {
            "dataset_id": dataset_id, "version": version_number, "schema": quality["columns"],
            "row_count": len(rows), "fingerprint": fingerprint, "source": source,
            "rows_path": rows_rel, "source_path": source_rel, "quality": quality, "created_at": _now(),
        }
        metadata["versions"].append(version)
        metadata["latest_version"] = version_number
        metadata["updated_at"] = _now()
        if task_id and task_id not in metadata["task_ids"]:
            metadata["task_ids"].append(task_id)
        self._save_metadata(metadata)
        return {**version, "dataset_id": dataset_id, "dataset_name": metadata["name"], "summary": self.summary(metadata)}

    def import_staged(self, token: str, *, filename: str | None = None, name: str = "", description: str = "", dataset_id: str | None = None, sheet: str | None = None, task_id: str | None = None) -> dict[str, Any]:
        safe = re.sub(r"[^a-f0-9]", "", token)
        candidates = list(self.staging.glob(f"{safe}.*")) if safe == token else []
        if len(candidates) != 1:
            raise ValueError("导入暂存文件不存在或已过期")
        path = candidates[0]
        try:
            original_filename = Path(filename or path.name).name
            return self.import_data(filename=original_filename, raw=path.read_bytes(), name=name, description=description, dataset_id=dataset_id, sheet=sheet, task_id=task_id)
        finally:
            path.unlink(missing_ok=True)

    def attach(self, dataset_id: str, task_id: str) -> dict[str, Any]:
        metadata = self._load_metadata(dataset_id)
        if task_id not in metadata.get("task_ids", []):
            metadata.setdefault("task_ids", []).append(task_id)
            metadata["updated_at"] = _now()
            self._save_metadata(metadata)
        return self.summary(metadata)


def dataset_service(settings: Settings) -> DatasetService:
    return DatasetService(settings)
